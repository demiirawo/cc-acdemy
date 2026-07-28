#!/usr/bin/env python3
"""Parse a FreeAgent company export (.xlsx) into SQL that upserts public.client_invoices.

    python3 scripts/parse_freeagent_export.py <export.xlsx> > /tmp/invoices.sql

The output is idempotent — it keys on (contact_organisation, reference,
invoice_date), so re-running with a newer export that overlaps the last one
updates those invoices in place instead of duplicating them. The refresh flow
is simply: download a new company export from FreeAgent, hand it to Claude (or
run this directly), apply the SQL.

The Invoices sheet is one row per invoice header followed by one row per line
item (line rows leave the contact/date columns blank), so lines attach to the
header row above them. Discounts, £18 Airtable add-ons, credits and one-offs
all live in those lines.
"""
import json
import sys
from datetime import date, datetime

import openpyxl


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def quote(s):
    if s is None:
        return "null"
    return "'" + str(s).replace("'", "''") + "'"


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Invoices" not in wb.sheetnames:
        sys.exit("No 'Invoices' sheet in that export")
    ws = wb["Invoices"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows)]
    col = {h: i for i, h in enumerate(headers)}

    def get(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    invoices = []
    current = None
    for row in rows:
        org = get(row, "Contact Organisation") or get(row, "Contact Name")
        d = iso(get(row, "Date"))
        if org and d:  # header row
            current = {
                "reference": str(get(row, "Reference") or "").strip(),
                "org": str(org).strip(),
                "date": d,
                "terms": int(num(get(row, "Payment Terms In Days")) or 0),
                "status": get(row, "Status") or None,
                "paid_date": iso(get(row, "Paid Date")),
                "paid_amount": num(get(row, "Paid Amount")),
                "net": num(get(row, "Net Amount")),
                "tax": num(get(row, "Sales Tax Amount")),
                "total": num(get(row, "Total Value")) or 0.0,
                "currency": get(row, "Currency") or "GBP",
                "lines": [],
            }
            invoices.append(current)
        item_type = get(row, "Item Type")
        if item_type and current is not None:
            current["lines"].append({
                "type": str(item_type),
                "description": str(get(row, "Description") or "").strip(),
                "quantity": num(get(row, "Quantity")),
                "price": num(get(row, "Price")),
            })

    usable = [i for i in invoices if i["reference"] and i["date"]]
    print(f"parsed {len(usable)} invoices from {path}", file=sys.stderr)

    export_name = path.split("/")[-1]
    values = []
    for i in usable:
        values.append(
            "({ref},{org},{d}::date,{terms},{status},{paid_d},{paid_a},{net},{tax},{total},{cur},{lines}::jsonb,{src})".format(
                ref=quote(i["reference"]),
                org=quote(i["org"]),
                d=quote(i["date"]),
                terms=i["terms"],
                status=quote(i["status"]),
                paid_d=f"{quote(i['paid_date'])}::date" if i["paid_date"] else "null",
                paid_a="null" if i["paid_amount"] is None else i["paid_amount"],
                net="null" if i["net"] is None else i["net"],
                tax="null" if i["tax"] is None else i["tax"],
                total=i["total"],
                cur=quote(i["currency"]),
                lines=quote(json.dumps(i["lines"])),
                src=quote(export_name),
            )
        )

    print("""insert into public.client_invoices
  (reference, contact_organisation, invoice_date, payment_terms_days, status,
   paid_date, paid_amount, net_amount, sales_tax_amount, total_value, currency, lines, source_export)
values
""" + ",\n".join(values) + """
on conflict (contact_organisation, reference, invoice_date) do update set
  payment_terms_days = excluded.payment_terms_days,
  status             = excluded.status,
  paid_date          = excluded.paid_date,
  paid_amount        = excluded.paid_amount,
  net_amount         = excluded.net_amount,
  sales_tax_amount   = excluded.sales_tax_amount,
  total_value        = excluded.total_value,
  currency           = excluded.currency,
  lines              = excluded.lines,
  source_export      = excluded.source_export,
  updated_at         = now();""")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("usage: parse_freeagent_export.py <export.xlsx>")
    main(sys.argv[1])
